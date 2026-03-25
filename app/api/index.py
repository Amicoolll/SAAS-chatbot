import logging
import os
from fastapi import APIRouter, BackgroundTasks, HTTPException, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import delete

from app.core.config import settings
from app.core.deps import get_tenant_user
from app.core.logging import log_operation
from app.db.session import get_db, SessionLocal
from app.db.models import Document, Chunk
from app.services import pipeline_state
from app.services.storage import list_files_recursive, read_text
from app.services.ingest.chunker import chunk_text
from app.services.openai_client import embed_texts

router = APIRouter(tags=["Indexing (pgvector)"])
logger = logging.getLogger(__name__)


def _read_indexable_text(path: str) -> str:
    if path.endswith(".txt") or path.endswith(".csv"):
        try:
            return read_text(path)
        except Exception as e:
            logger.warning(
                "Skipping text/CSV (unreadable) path=%s error=%s",
                path,
                type(e).__name__,
            )
            return ""
    if path.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail="XLSX indexing requires 'openpyxl'. Install dependencies and retry.",
            ) from e
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                lines: list[str] = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        if any(x.strip() for x in cells):
                            lines.append("\t".join(cells))
                return "\n".join(lines).strip()
            finally:
                wb.close()
        except Exception as e:
            # Password-protected / encrypted workbooks, corrupt files, etc.
            logger.warning(
                "Skipping XLSX (encrypted, corrupt, or unreadable) path=%s error=%s",
                path,
                type(e).__name__,
            )
            return ""
    if path.endswith(".pdf"):
        try:
            from pypdf import PdfReader
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail="PDF indexing requires 'pypdf'. Install dependencies and retry.",
            ) from e
        try:
            reader = PdfReader(path)
            pages: list[str] = []
            for page in reader.pages:
                pages.append(page.extract_text() or "")
            return "\n".join(pages).strip()
        except Exception as e:
            # Encrypted PDFs (cannot decrypt without password), corrupt files, etc.
            logger.warning(
                "Skipping PDF (encrypted or unreadable — not decrypted) path=%s error=%s",
                path,
                type(e).__name__,
            )
            return ""
    return ""


def _run_index(
    db: Session,
    tenant_id: str,
    user_id: str,
    max_files: int,
) -> dict:
    """Index raw files into Postgres + pgvector. Used by route and background task."""
    base_dir = os.path.join("data", f"user_{user_id}")
    raw_dir = os.path.join(base_dir, "raw")

    if not os.path.exists(raw_dir):
        raise HTTPException(status_code=400, detail="No raw files found. Run /drive/sync first.")

    raw_files = list_files_recursive(raw_dir)
    raw_files = [
        p
        for p in raw_files
        if p.endswith(".txt")
        or p.endswith(".csv")
        or p.endswith(".pdf")
        or p.endswith(".xlsx")
    ][:max_files]
    total_files = len(raw_files)

    pipeline_state.update_index_progress(
        tenant_id,
        user_id,
        phase="embedding",
        current=0,
        total=total_files,
        current_file=None,
        chunks_so_far=0,
    )

    batch_size = settings.EMBED_BATCH_SIZE
    chunk_size = settings.CHUNK_SIZE
    chunk_overlap = settings.CHUNK_OVERLAP

    docs_indexed = 0
    chunks_indexed = 0
    files_skipped_unreadable = 0
    files_skipped_error = 0

    def _finish_file_progress() -> None:
        pipeline_state.update_index_progress(
            tenant_id,
            user_id,
            phase="embedding",
            current=file_idx,
            total=total_files,
            current_file=name,
            chunks_so_far=chunks_indexed,
        )

    for file_idx, path in enumerate(raw_files, start=1):
        name = os.path.basename(path)
        pipeline_state.update_index_progress(
            tenant_id,
            user_id,
            phase="embedding",
            current=file_idx - 1,
            total=total_files,
            current_file=name,
            chunks_so_far=chunks_indexed,
        )
        if path.endswith(".txt"):
            mime_type = "text/plain"
        elif path.endswith(".csv"):
            mime_type = "text/csv"
        elif path.endswith(".xlsx"):
            mime_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            mime_type = "application/pdf"

        try:
            text = _read_indexable_text(path)
            # Postgres/psycopg reject NUL in text parameters; PDFs often contain them.
            text = text.replace("\x00", "")
            if not text.strip():
                files_skipped_unreadable += 1
                logger.info(
                    "Skipping file (empty or unreadable / not decrypted): %s", path
                )
                _finish_file_progress()
                continue
            chunks = chunk_text(text, chunk_size=chunk_size, overlap=chunk_overlap)
            if not chunks:
                _finish_file_progress()
                continue

            embeddings: list[list[float]] = []
            for batch_start in range(0, len(chunks), batch_size):
                embeddings.extend(
                    embed_texts(chunks[batch_start : batch_start + batch_size])
                )

            if len(embeddings) != len(chunks):
                logger.warning(
                    "Skipping file (embedding count mismatch) path=%s chunks=%s embeddings=%s",
                    path,
                    len(chunks),
                    len(embeddings),
                )
                files_skipped_error += 1
                _finish_file_progress()
                continue
            bad_dim = next(
                (len(e) for e in embeddings if len(e) != settings.EMBED_DIM),
                None,
            )
            if bad_dim is not None:
                logger.warning(
                    "Skipping file (embedding dimension mismatch) path=%s got=%s expected=%s model=%s",
                    path,
                    bad_dim,
                    settings.EMBED_DIM,
                    settings.OPENAI_EMBEDDING_MODEL,
                )
                files_skipped_error += 1
                _finish_file_progress()
                continue

            drive_file_id = f"local::{user_id}::{name}"

            doc = db.query(Document).filter(
                Document.tenant_id == tenant_id,
                Document.user_id == user_id,
                Document.drive_file_id == drive_file_id,
            ).first()

            if not doc:
                doc = Document(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    drive_file_id=drive_file_id,
                    name=name,
                    mime_type=mime_type,
                    modified_time="",
                    web_view_link="",
                )
                db.add(doc)
                db.commit()
                db.refresh(doc)
            else:
                db.execute(delete(Chunk).where(Chunk.document_id == doc.id))
                db.commit()

            for idx, (c, e) in enumerate(zip(chunks, embeddings)):
                db.add(Chunk(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    document_id=doc.id,
                    chunk_index=idx,
                    content=c,
                    embedding=e,
                ))

            db.commit()
            docs_indexed += 1
            chunks_indexed += len(chunks)
            _finish_file_progress()
        except Exception:
            db.rollback()
            files_skipped_error += 1
            logger.exception("index_file_failed path=%s", path)
            _finish_file_progress()

    log_operation(
        logger, "index_run", tenant_id=tenant_id, user_id=user_id,
        docs_indexed=docs_indexed, chunks_indexed=chunks_indexed,
        files_skipped_unreadable=files_skipped_unreadable,
        files_skipped_error=files_skipped_error,
    )
    return {
        "tenant_id": tenant_id,
        "user_id": user_id,
        "docs_indexed": docs_indexed,
        "chunks_indexed": chunks_indexed,
        "total_files_planned": total_files,
        "files_skipped_unreadable": files_skipped_unreadable,
        "files_skipped_error": files_skipped_error,
        "note": "Indexed raw/ files into Postgres + pgvector. Skipped encrypted/unreadable/empty files.",
    }


def _index_background_task(tenant_id: str, user_id: str, max_files: int) -> None:
    db = SessionLocal()
    pipeline_state.mark_index_running(tenant_id, user_id)
    try:
        result = _run_index(db, tenant_id, user_id, max_files)
        pipeline_state.mark_index_success(tenant_id, user_id, result)
    except HTTPException as e:
        detail = e.detail if isinstance(e.detail, str) else str(e.detail)
        pipeline_state.mark_index_error(tenant_id, user_id, detail)
    except Exception as e:
        logger.exception("index_run_background_failed tenant_id=%s user_id=%s", tenant_id, user_id)
        pipeline_state.mark_index_error(tenant_id, user_id, str(e))
    finally:
        db.close()


@router.post("/index/run")
def index_run(
    background_tasks: BackgroundTasks,
    tenant_user: tuple[str, str] = Depends(get_tenant_user),
    max_files: int = Query(
        100,
        ge=1,
        le=5000,
        description="Max raw .txt/.csv files to embed this run (OpenAI cost scales with chunks). Use 385+ when ready; progress total matches this batch.",
    ),
    background: bool = True,
    db: Session = Depends(get_db),
):
    """Index raw folder into pgvector. If background=True (default), returns 202 and runs in background.

    Poll **GET /pipeline/status** for `index.progress` (`current` / `total` / `current_file` / `chunks_so_far` / `percent`).
    """
    tenant_id, user_id = tenant_user

    base_dir = os.path.join("data", f"user_{user_id}")
    raw_dir = os.path.join(base_dir, "raw")
    if not os.path.exists(raw_dir):
        raise HTTPException(status_code=400, detail="No raw files found. Run /drive/sync first.")

    if background:
        background_tasks.add_task(_index_background_task, tenant_id, user_id, max_files)
        return {
            "status": "accepted",
            "message": "Indexing started in background. Poll GET /pipeline/status until index is not running.",
            "tenant_id": tenant_id,
            "user_id": user_id,
        }
    pipeline_state.mark_index_running(tenant_id, user_id)
    try:
        result = _run_index(db, tenant_id, user_id, max_files)
        pipeline_state.mark_index_success(tenant_id, user_id, result)
        return result
    except HTTPException as e:
        detail = e.detail if isinstance(e.detail, str) else str(e.detail)
        pipeline_state.mark_index_error(tenant_id, user_id, detail)
        raise
    except Exception as e:
        pipeline_state.mark_index_error(tenant_id, user_id, str(e))
        raise
