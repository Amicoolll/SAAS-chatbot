#!/usr/bin/env python3
"""Rewrite expected label columns in a query-understanding .xlsx using analyze_query()."""

from __future__ import annotations

import argparse
import importlib.util
import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_root))

import openpyxl

from app.services.query_understanding import analyze_query

_csv_script = _root / "scripts" / "compare_query_expected_csv.py"
_spec = importlib.util.spec_from_file_location("qu_expected_csv", _csv_script)
_qu = importlib.util.module_from_spec(_spec)
assert _spec.loader
_spec.loader.exec_module(_qu)
KEYS = _qu.KEYS
_settings = _qu._settings


def main() -> int:
    p = argparse.ArgumentParser(
        description="Fill domain/intent/... columns from analyze_query (matches scripts/golden_expected_outputs behavior).",
    )
    p.add_argument("xlsx_path", type=Path)
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output path (default: overwrite input).",
    )
    args = p.parse_args()

    path = args.xlsx_path
    if not path.is_file():
        print(f"Not a file: {path}", file=sys.stderr)
        return 2

    out = args.output or path
    settings = _settings()

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        print("No data rows.", file=sys.stderr)
        return 2

    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        h = (cell.value if cell.value is not None else None)
        if h is not None:
            headers[str(h).strip()] = col

    if "query" not in headers:
        print("Sheet needs a 'query' column.", file=sys.stderr)
        return 2
    missing = [k for k in KEYS if k not in headers]
    if missing:
        print(f"Sheet missing columns: {missing}", file=sys.stderr)
        return 2

    q_col = headers["query"]
    updated = 0
    for row in range(2, ws.max_row + 1):
        q_cell = ws.cell(row=row, column=q_col)
        q = q_cell.value
        if q is None or not str(q).strip():
            continue
        result = analyze_query(str(q).strip(), settings=settings).model_dump()
        for key in KEYS:
            col = headers[key]
            dest = ws.cell(row=row, column=col)
            val = result[key]
            if isinstance(val, bool):
                dest.value = val
            else:
                dest.value = val
        updated += 1

    wb.save(out)
    print(f"Updated {updated} rows → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
